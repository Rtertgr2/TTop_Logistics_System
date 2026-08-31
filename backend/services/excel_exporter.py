import io

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


def generate_all_routes_manifest_excel(routes: list[dict]) -> bytes:
    """สร้างไฟล์ Excel รายงานการจัดส่งรวมทุกคัน พร้อมสร้าง Sheet แยกสำหรับคนขับแต่ละคัน"""
    wb = openpyxl.Workbook()

    # ── Sheet 1: สรุปภาพรวมการจัดขนส่ง ─────────────────────────────
    ws_summary = wb.active
    ws_summary.title = "ภาพรวมการจัดขนส่ง"
    ws_summary.views.sheetView[0].showGridLines = True

    # Styling
    header_font = Font(name="Cordia New", size=18, bold=True, color="1E293B")
    sub_font = Font(name="Cordia New", size=14, bold=True, color="475569")
    table_header_font = Font(name="Cordia New", size=14, bold=True, color="FFFFFF")
    table_header_fill = PatternFill(
        start_color="1E40AF", end_color="1E40AF", fill_type="solid"
    )

    border_thin = Side(style="thin", color="CBD5E1")
    cell_border = Border(
        left=border_thin, right=border_thin, top=border_thin, bottom=border_thin
    )

    # Header Section
    ws_summary.cell(
        row=1, column=1, value="🚛 รายงานสรุปแผนการจัดขนส่งสินค้าประจำวัน"
    ).font = header_font
    ws_summary.cell(
        row=2,
        column=1,
        value=f"จำนวนรถที่ใช้งาน: {len(routes)} คัน | รวมจุดส่งทั้งหมด: {sum(len(r.get('stops', [])) for r in routes)} จุด",
    ).font = sub_font

    # Summary Table Headers
    sum_headers = [
        "ลำดับ",
        "คันที่",
        "ทะเบียนรถ",
        "พนักงานขับรถ",
        "ความจุสูงสุด (kg)",
        "น้ำหนักบรรทุกจริง (kg)",
        "% ความจุ",
        "จำนวนจุดส่ง",
    ]
    ws_summary.append([])  # Row 3 blank

    ws_summary.row_dimensions[4].height = 28
    for col_num, h_text in enumerate(sum_headers, 1):
        cell = ws_summary.cell(row=4, column=col_num, value=h_text)
        cell.font = table_header_font
        cell.fill = table_header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = cell_border

    # Summary Data Rows
    total_all_weight = 0
    total_all_stops = 0

    for idx, r in enumerate(routes, 1):
        stops = r.get("stops", [])
        w = r.get("total_weight", 0)
        cap = r.get("capacity", 3750)
        pct = (w / cap * 100) if cap else 0

        total_all_weight += w
        total_all_stops += len(stops)

        row_data = [
            idx,
            r.get("name", f"รถคันที่ {idx}"),
            r.get("plate", "-"),
            r.get("driver", "-"),
            cap,
            w,
            f"{pct:.1f}%",
            len(stops),
        ]

        row_idx = 4 + idx
        ws_summary.row_dimensions[row_idx].height = 22
        for c_idx, val in enumerate(row_data, 1):
            cell = ws_summary.cell(row=row_idx, column=c_idx, value=val)
            cell.font = Font(name="Cordia New", size=13)
            cell.border = cell_border
            if c_idx in [1, 5, 6, 7, 8]:
                cell.alignment = Alignment(horizontal="center", vertical="center")

    # ── Sheet 2..N: ใบสั่งงานคนขับรถแต่ละคัน (Driver Manifest) ────
    for idx, r in enumerate(routes, 1):
        v_name = r.get("name", f"รถคันที่ {idx}")
        sheet_title = f"รถคันที่ {idx} ({r.get('plate', 'คิวงาน')})"[:30]

        ws = wb.create_sheet(title=sheet_title)
        ws.views.sheetView[0].showGridLines = True

        # Title
        ws.cell(
            row=1,
            column=1,
            value=f"📋 ใบสั่งงานจัดส่งสินค้า — {v_name} (ทะเบียน {r.get('plate', '-')})",
        ).font = Font(name="Cordia New", size=18, bold=True, color="1E3A8A")
        ws.cell(
            row=2,
            column=1,
            value=f"พนักงานขับรถ: {r.get('driver', '-')} | ความจุรถ: {r.get('capacity', 0)} kg | น้ำหนักจัดส่งรวม: {r.get('total_weight', 0)} kg",
        ).font = sub_font

        # Manifest Headers
        man_headers = [
            "ลำดับจุดส่ง",
            "เลขที่ SO",
            "ชื่อลูกค้า / ร้านค้า",
            "สถานที่จัดส่งสินค้า",
            "เขต / โซน",
            "น้ำหนัก (kg)",
            "ลายเซ็นผู้รับสินค้า",
        ]
        ws.append([])  # Row 3

        ws.row_dimensions[4].height = 26
        for c_idx, h_text in enumerate(man_headers, 1):
            cell = ws.cell(row=4, column=c_idx, value=h_text)
            cell.font = table_header_font
            cell.fill = PatternFill(
                start_color="1E3A8A", end_color="1E3A8A", fill_type="solid"
            )
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = cell_border

        # Manifest Data Rows
        stops = r.get("stops", [])
        for s_idx, stop in enumerate(stops, 1):
            r_idx = 4 + s_idx
            ws.row_dimensions[r_idx].height = 24

            row_vals = [
                s_idx,
                stop.get("order_number", "-"),
                stop.get("customer", "-"),
                stop.get("address", "-"),
                stop.get("zone", "-"),
                stop.get("weight", 0),
                "",  # Signature box
            ]

            for c_idx, val in enumerate(row_vals, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=val)
                cell.font = Font(name="Cordia New", size=13)
                cell.border = cell_border
                if c_idx in [1, 2, 5, 6]:
                    cell.alignment = Alignment(horizontal="center", vertical="center")

        # เพิ่ม Sheet สำหรับรายละเอียดสินค้า (OrderItem Details)
        detail_sheet_title = f"รายละเอียดสินค้า ({r.get('plate', 'คิวงาน')})"[:30]
        ws_detail = wb.create_sheet(title=detail_sheet_title)
        ws_detail.views.sheetView[0].showGridLines = True

        # Detail Title
        ws_detail.cell(
            row=1,
            column=1,
            value=f"📦 รายละเอียดสินค้า — {v_name} (ทะเบียน {r.get('plate', '-')})",
        ).font = Font(name="Cordia New", size=18, bold=True, color="1E3A8A")
        ws_detail.cell(
            row=2, column=1, value=f"น้ำหนักรวม: {r.get('total_weight', 0)} kg"
        ).font = sub_font

        # Detail Headers
        detail_headers = [
            "ลำดับ",
            "เลขที่ SO",
            "ชื่อลูกค้า",
            "รหัสสินค้า",
            "ชื่อสินค้า",
            "จำนวน",
            "หน่วย",
            "น้ำหนักต่อหน่วย (kg)",
            "น้ำหนักรวม (kg)",
        ]
        ws_detail.append([])  # Row 3

        ws_detail.row_dimensions[4].height = 26
        for c_idx, h_text in enumerate(detail_headers, 1):
            cell = ws_detail.cell(row=4, column=c_idx, value=h_text)
            cell.font = table_header_font
            cell.fill = PatternFill(
                start_color="1E3A8A", end_color="1E3A8A", fill_type="solid"
            )
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = cell_border

        # Detail Data Rows
        detail_row = 5
        for s_idx, stop in enumerate(stops, 1):
            products = stop.get("products", [])
            if not products:
                # ถ้าไม่มี products ให้แสดงแค่ order info
                ws_detail.cell(row=detail_row, column=1, value=s_idx).font = Font(
                    name="Cordia New", size=13
                )
                ws_detail.cell(
                    row=detail_row, column=2, value=stop.get("order_number", "-")
                ).font = Font(name="Cordia New", size=13)
                ws_detail.cell(
                    row=detail_row, column=3, value=stop.get("customer", "-")
                ).font = Font(name="Cordia New", size=13)
                ws_detail.cell(row=detail_row, column=4, value="-").font = Font(
                    name="Cordia New", size=13
                )
                ws_detail.cell(row=detail_row, column=5, value="-").font = Font(
                    name="Cordia New", size=13
                )
                ws_detail.cell(row=detail_row, column=6, value=0).font = Font(
                    name="Cordia New", size=13
                )
                ws_detail.cell(row=detail_row, column=7, value="-").font = Font(
                    name="Cordia New", size=13
                )
                ws_detail.cell(row=detail_row, column=8, value=0).font = Font(
                    name="Cordia New", size=13
                )
                ws_detail.cell(
                    row=detail_row, column=9, value=stop.get("weight", 0)
                ).font = Font(name="Cordia New", size=13)
                detail_row += 1
            else:
                # แสดงแต่ละ product ใน order
                for p_idx, product in enumerate(products):
                    ws_detail.cell(
                        row=detail_row, column=1, value=s_idx if p_idx == 0 else ""
                    ).font = Font(name="Cordia New", size=13)
                    ws_detail.cell(
                        row=detail_row,
                        column=2,
                        value=stop.get("order_number", "-") if p_idx == 0 else "",
                    ).font = Font(name="Cordia New", size=13)
                    ws_detail.cell(
                        row=detail_row,
                        column=3,
                        value=stop.get("customer", "-") if p_idx == 0 else "",
                    ).font = Font(name="Cordia New", size=13)
                    ws_detail.cell(
                        row=detail_row, column=4, value=product.get("code", "-")
                    ).font = Font(name="Cordia New", size=13)
                    ws_detail.cell(
                        row=detail_row, column=5, value=product.get("name", "-")
                    ).font = Font(name="Cordia New", size=13)
                    ws_detail.cell(
                        row=detail_row, column=6, value=product.get("quantity", 0)
                    ).font = Font(name="Cordia New", size=13)
                    ws_detail.cell(
                        row=detail_row, column=7, value=product.get("unit", "-")
                    ).font = Font(name="Cordia New", size=13)
                    ws_detail.cell(
                        row=detail_row, column=8, value=product.get("product_weight", 0)
                    ).font = Font(name="Cordia New", size=13)
                    ws_detail.cell(
                        row=detail_row, column=9, value=product.get("item_weight", 0)
                    ).font = Font(name="Cordia New", size=13)
                    detail_row += 1

        # Auto column width for detail sheet
        for col in ws_detail.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws_detail.column_dimensions[col_letter].width = max(max_len + 4, 12)

        # Auto column width adjustment
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    # Auto column width for summary
    for col in ws_summary.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws_summary.column_dimensions[col_letter].width = max(max_len + 4, 14)

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()
